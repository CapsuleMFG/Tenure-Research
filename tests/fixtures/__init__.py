"""Builders for small synthetic XLSX fixtures used by the cleaner tests.

These mimic the shapes of real ERS files (wide monthly time series and
WASDE-style year+quarter tables) without shipping any real USDA data.
The :func:`build_all` function is invoked from ``conftest.py`` so that
fixtures are regenerated whenever the test suite runs — they are not
checked into git.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).resolve().parent


def build_wide_format(path: Path) -> None:
    """Create a livestock-prices-style fixture: 4 header rows, dates in col A."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Historical"

    ws["A1"] = None
    ws["B1"] = "Cattle prices"
    ws["E1"] = "Hog prices"

    ws["A2"] = None
    ws["B2"] = "TX/OK/NM steers"
    ws["C2"] = "Feeder steers 500-550 lbs"
    ws["D2"] = "Feeder steers 750-800 lbs"
    ws["E2"] = "Barrows and gilts"

    ws["A3"] = None
    ws["B3"] = "35-65 percent Choice"
    ws["C3"] = "Medium and large"
    ws["D3"] = "Medium and large"
    ws["E3"] = "National base 51-52 percent"

    ws["A4"] = "Period"
    ws["B4"] = "Dollars/cwt"
    ws["C4"] = "Dollars/cwt"
    ws["D4"] = "Dollars/cwt"
    ws["E4"] = "Dollars/cwt"

    sample = [
        (datetime(2024, 1, 1), 175.10, 250.50, 200.30, "NA"),
        (datetime(2024, 2, 1), 180.45, 255.10, 205.80, "NA"),
        (datetime(2024, 3, 1), 185.20, 260.75, 210.40, 75.55),
        (datetime(2024, 4, 1), None, 265.00, 215.10, 78.90),     # truly blank value
        (datetime(2024, 5, 1), 192.10, 270.25, 220.60, 80.20),
        (None, None, None, None, None),                            # blank separator
        (datetime(2024, 6, 1), "  ", 275.50, 225.30, "--"),       # whitespace + dash null
        (datetime(2024, 7, 1), 198.40, 280.10, 230.00, 82.50),
    ]
    for row_idx, row in enumerate(sample, start=5):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_wasde_format(path: Path) -> None:
    """Create a WASDE-style fixture: 3 header rows, year (sparse) + quarter cols."""
    wb = Workbook()
    ws = wb.active
    ws.title = "WASDE_Beef"

    ws["A1"] = "Beef: Supply and disappearance"
    ws["A2"] = "Year and qtr"
    ws["C2"] = "Production"
    ws["A3"] = None
    ws["B3"] = None
    ws["C3"] = "Commercial"
    ws["D3"] = "Farm"
    ws["E3"] = "Total"
    ws["L3"] = "Total disappearance"

    sample = [
        (2023, "Q1 Jan-Mar", 6800.0, 25.0, 6825.0, None, None, None, None, None, None, 7100.0),
        (None, "Q2 Apr-Jun", 6700.0, 8.5, 6708.5, None, None, None, None, None, None, 6900.0),
        (None, "Q3 Jul-Sep", 6650.0, 8.5, 6658.5, None, None, None, None, None, None, 6850.0),
        (None, "Q4 Oct-Dec", 6800.0, 25.5, 6825.5, None, None, None, None, None, None, 6950.0),
        (None, "Yr Jan-Dec", 26950.0, 67.5, 27017.5, None, None, None, None, None, None, 27800.0),
        (2024, "Q1 Jan-Mar", 6550.0, 25.0, 6575.0, None, None, None, None, None, None, 7120.0),
        (None, "Q2 Apr-Jun", 6760.0, 8.0, 6768.0, None, None, None, None, None, None, 7040.0),
        # A row with explicit "NA" for production — should round-trip as null
        (None, "Q3 Jul-Sep", "NA", 8.0, 6700.0, None, None, None, None, None, None, 7000.0),
    ]
    for row_idx, row in enumerate(sample, start=4):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_all(out_dir: Path = FIXTURES_DIR) -> dict[str, Path]:
    """Build every fixture under ``out_dir`` and return their paths."""
    paths = {
        "wide_format": out_dir / "wide_format.xlsx",
        "wasde_format": out_dir / "wasde_format.xlsx",
    }
    build_wide_format(paths["wide_format"])
    build_wasde_format(paths["wasde_format"])
    return paths
