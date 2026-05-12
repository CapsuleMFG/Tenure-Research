"""Read raw ERS XLSX files and emit tidy ``observations.parquet``.

The cleaning logic is intentionally driven by the catalog: each
:class:`~usda_sandbox.catalog.SeriesDefinition` describes the sheet, the
header offset, the date column(s), and the value column for one series.
Two raw layouts are supported:

* **Wide time-series** (e.g. ``livestock-prices.xlsx`` Historical, ``wholesale-prices.xlsx``
  Historical) — column A contains a real date cell, column letters B-onward
  contain values. The catalog encodes this with ``date_column = "A"``.
* **WASDE quarterly** (e.g. ``meat-supply-and-disappearance...xlsx`` ``WASDE_*``)
  — year lives in one (sparse, forward-filled) column and the quarter label
  ("Q1 Jan-Mar", "Yr Jan-Dec", ...) in another. The catalog encodes this with
  ``date_column = "A+B"``. ``"Yr Jan-Dec"`` rollup rows are dropped.
"""

from __future__ import annotations

import calendar
from collections.abc import Iterable
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .catalog import SeriesDefinition, load_catalog
from .futures import append_futures_to_observations

OBSERVATIONS_SCHEMA: dict[str, Any] = {
    "series_id": pl.Utf8,
    "series_name": pl.Utf8,
    "commodity": pl.Utf8,
    "metric": pl.Utf8,
    "unit": pl.Utf8,
    "frequency": pl.Utf8,
    "period_start": pl.Date,
    "period_end": pl.Date,
    "value": pl.Float64,
    "source_file": pl.Utf8,
    "source_sheet": pl.Utf8,
    "ingested_at": pl.Datetime("us", "UTC"),
}

_QUARTER_BOUNDS: dict[str, tuple[int, int]] = {
    "Q1": (1, 3),
    "Q2": (4, 6),
    "Q3": (7, 9),
    "Q4": (10, 12),
}

_NULL_TOKENS = frozenset({"", "NA", "N/A", "n/a", "-", "--", ".", "(D)", "(NA)"})


def _coerce_float(value: Any) -> float | None:
    """Convert an XLSX cell into a float, or ``None`` for known null markers."""
    if value is None:
        return None
    if isinstance(value, bool):
        # Surprising but openpyxl can return bool for "TRUE"/"FALSE" cells; reject.
        return None
    if isinstance(value, int | float):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if s in _NULL_TOKENS:
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def _coerce_year(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        try:
            year = int(value)
        except (TypeError, ValueError):
            return None
    elif isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            year = int(float(s))
        except ValueError:
            return None
    else:
        return None
    if 1900 <= year <= 2100:
        return year
    return None


def _month_bounds(d: date) -> tuple[date, date]:
    start = d.replace(day=1)
    last_day = calendar.monthrange(start.year, start.month)[1]
    return start, start.replace(day=last_day)


def _quarter_bounds(year: int, label: str) -> tuple[date, date] | None:
    """Map ``"Q1 Jan-Mar"``-style labels to a (start, end) pair."""
    code = label.strip().split()[0].upper() if label and label.strip() else ""
    if code not in _QUARTER_BOUNDS:
        return None
    sm, em = _QUARTER_BOUNDS[code]
    last_day = calendar.monthrange(year, em)[1]
    return date(year, sm, 1), date(year, em, last_day)


def _annual_bounds(d: date) -> tuple[date, date]:
    return date(d.year, 1, 1), date(d.year, 12, 31)


def _to_python_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.strip()).date()
        except ValueError:
            return None
    return None


def _is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(
        cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row
    )


def _column_index(letter: str) -> int:
    """1-based column index from a column letter."""
    return int(column_index_from_string(letter.strip().upper()))


def _resolve_date_columns(spec: str) -> tuple[int, int | None]:
    """Return ``(date_or_year_idx, quarter_idx_or_None)`` from a date_column spec."""
    parts = [piece.strip() for piece in spec.split("+") if piece.strip()]
    if len(parts) == 1:
        return _column_index(parts[0]), None
    if len(parts) == 2:
        return _column_index(parts[0]), _column_index(parts[1])
    raise ValueError(f"Unsupported date_column spec: {spec!r}")


def _row_to_observation(
    series_def: SeriesDefinition,
    period_start: date,
    period_end: date,
    value: float | None,
    ingested_at: datetime,
) -> dict[str, Any]:
    return {
        "series_id": series_def.series_id,
        "series_name": series_def.series_name,
        "commodity": series_def.commodity,
        "metric": series_def.metric,
        "unit": series_def.unit,
        "frequency": series_def.frequency,
        "period_start": period_start,
        "period_end": period_end,
        "value": value,
        "source_file": series_def.source_file,
        "source_sheet": series_def.source_sheet,
        "ingested_at": ingested_at,
    }


def _empty_frame() -> pl.DataFrame:
    return pl.DataFrame(schema=OBSERVATIONS_SCHEMA)


def _frame_from_rows(rows: Iterable[dict[str, Any]]) -> pl.DataFrame:
    rows = list(rows)
    if not rows:
        return _empty_frame()
    df = pl.DataFrame(rows)
    # Ensure stable column order + dtypes regardless of which cells were present.
    return df.select(
        [pl.col(name).cast(dtype) for name, dtype in OBSERVATIONS_SCHEMA.items()]
    )


def clean_series(series_def: SeriesDefinition, raw_path: Path | str) -> pl.DataFrame:
    """Extract a single series from a raw XLSX file as a tidy DataFrame.

    Rows whose value cell is a known null marker (``NA``, ``--`` etc.) are kept
    with ``value = None`` so callers can see coverage gaps. Rows whose date
    cannot be parsed (blank rows, group separators, sub-headers, annual
    rollups) are silently dropped.
    """
    raw_path = Path(raw_path)
    date_idx, quarter_idx = _resolve_date_columns(series_def.date_column)
    if not series_def.value_columns:
        raise ValueError(
            f"Series {series_def.series_id!r} has no value_columns configured"
        )
    value_idx = _column_index(series_def.value_columns[0])

    wb = load_workbook(raw_path, read_only=True, data_only=True)
    ingested_at = datetime.now(UTC)
    rows: list[dict[str, Any]] = []
    last_year: int | None = None

    try:
        if series_def.source_sheet not in wb.sheetnames:
            raise KeyError(
                f"Sheet {series_def.source_sheet!r} not found in {raw_path.name}; "
                f"available: {wb.sheetnames}"
            )
        ws = wb[series_def.source_sheet]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx <= series_def.header_rows_to_skip:
                continue
            if _is_blank_row(row):
                continue

            value_cell = row[value_idx - 1] if value_idx - 1 < len(row) else None

            if quarter_idx is not None:
                year_cell = row[date_idx - 1] if date_idx - 1 < len(row) else None
                qtr_cell = row[quarter_idx - 1] if quarter_idx - 1 < len(row) else None
                year = _coerce_year(year_cell)
                if year is not None:
                    last_year = year
                if last_year is None or qtr_cell is None:
                    continue
                qtr_label = str(qtr_cell).strip()
                # Skip annual rollup rows when we're collecting quarterly data
                if qtr_label.lower().startswith("yr"):
                    continue
                bounds = _quarter_bounds(last_year, qtr_label)
            else:
                date_cell = row[date_idx - 1] if date_idx - 1 < len(row) else None
                py_date = _to_python_date(date_cell)
                if py_date is None:
                    continue
                if series_def.frequency == "monthly":
                    bounds = _month_bounds(py_date)
                elif series_def.frequency == "annual":
                    bounds = _annual_bounds(py_date)
                else:
                    bounds = (py_date, py_date)

            if bounds is None:
                continue

            value = _coerce_float(value_cell)
            rows.append(
                _row_to_observation(
                    series_def,
                    bounds[0],
                    bounds[1],
                    value,
                    ingested_at,
                )
            )
    finally:
        wb.close()

    return _frame_from_rows(rows)


def clean_all(
    catalog_path: Path | str,
    raw_dir: Path | str,
    out_path: Path | str,
) -> pl.DataFrame:
    """Process every catalog entry and write a single tidy parquet file.

    XLSX-derived entries are parsed inline. Futures-derived entries
    (``source_file.startswith("futures:")``) are skipped here; the futures
    module's :func:`append_futures_to_observations` handles them after
    the XLSX pass completes.
    """
    catalog = load_catalog(catalog_path)
    raw_dir = Path(raw_dir)
    frames: list[pl.DataFrame] = []
    has_futures = False
    for series_def in catalog:
        if series_def.source_file.startswith("futures:"):
            has_futures = True
            continue
        raw_path = raw_dir / series_def.source_file
        if not raw_path.exists():
            print(
                f"[clean_all] skipping {series_def.series_id}: missing {raw_path}",
                flush=True,
            )
            continue
        frames.append(clean_series(series_def, raw_path))

    if not frames:
        combined = _empty_frame()
    else:
        combined = pl.concat(frames, how="vertical_relaxed").unique(
            subset=["series_id", "period_start", "source_file"], keep="first"
        )
        combined = combined.sort(["series_id", "period_start"])

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    combined.write_parquet(out)

    if has_futures:
        append_futures_to_observations(obs_path=out)
        combined = pl.read_parquet(out)

    return combined
